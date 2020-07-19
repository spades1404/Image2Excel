from PIL import Image
import webcolors
from matplotlib import colors as mplcol
from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

#############################################################################
imageName = "test.jpg" #place your image in here e.g picture.jpg                   ##
outputSheet = "output.xlsx" #you can change the name of this               ##
#############################################################################


im = Image.open(imageName) #grab image
pic = im.load() #load image
size = im.size #get image size

#def width and height of image
width = size[0]
height = size[1]


def rgb2hex(rgb):
    return webcolors.rgb_to_hex(rgb).upper()[1:] + "00" #converts to rgb to hex (added alpha channel is opaque)

def hex2patterfillobj(hex):
    return PatternFill(
        start_color= hex,
        end_color= hex,
        fill_type= "solid"
    )



wb = Workbook()
ws = wb.active

for i in range(height):
    for j in range(width):
        ws.cell(row = i+1, column = j+1).fill = (hex2patterfillobj(rgb2hex(pic[j,i])))
    print("Row done")


wb.save(outputSheet)

